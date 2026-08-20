import streamlit as st
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="出貨自動化工具箱", page_icon="📦", layout="wide")

st.title("📦 訂單處理與出貨自動化工具箱")

tab1, tab2 = st.tabs(["📋 工具一：訂單自動分類與整理", "🚚 工具二：轉 7-11 批量匯入檔"])

COLOR_PALETTE = [
    "FFF2CC", "D9EAD3", "C9DAF8", "FCE5CD", "EAD1DC", "D9D2E9", "E0F7FA", "F3E5F5"
]

def parse_taiwan_address(address):
    if not isinstance(address, str) or not address:
        return "", "", "", ""
    address = address.strip()
    address = re.sub(r'[\,\s]*台灣[\,\s]*$', '', address, flags=re.IGNORECASE)
    address = re.sub(r'^\d{3,5}\s*', '', address.strip())
    address = address.replace("台", "臺")
    
    city, district, road, rest = "", "", "", ""
    city_match = re.search(r'(.+?[縣市])', address)
    if city_match:
        city = city_match.group(1)
        address = address[len(city):]
    
    dist_match = re.search(r'(.+?[區鎮鄉市])', address)
    if dist_match:
        district = dist_match.group(1)
        address = address[len(district):]
        
    road_match = re.search(r'(.+?[路街大道(段)])', address)
    if road_match:
        road = road_match.group(1)
        rest = address[len(road):]
    else:
        rest = address
        
    return city, district, road, rest

def parse_store_info(store_str):
    if not isinstance(store_str, str) or not store_str:
        return "", ""
    store_str = store_str.strip()
    code_match = re.search(r'\d{5,8}', store_str)
    store_code = code_match.group(0) if code_match else ""
    
    store_name = re.sub(r'\[.*?\]', '', store_str)
    store_name = re.sub(r'\(.*?\)', '', store_name)
    store_name = re.sub(r'（.*?）', '', store_name)
    store_name = re.sub(r'代碼[\:：]?\d+', '', store_name)
    store_name = store_name.replace("台", "臺").strip()
    return store_code, store_name

def format_phone_number(phone_val):
    if pd.isna(phone_val) or phone_val is None:
        return ""
    phone_str = str(phone_val).split('.')[0].strip()
    phone_str = re.sub(r'[^\d]', '', phone_str)
    if len(phone_str) == 9 and phone_str.startswith('9'):
        phone_str = '0' + phone_str
    return phone_str

def clean_overseas_columns(df_in, col_phone_name):
    df_out = df_in.copy()
    cols_to_drop = [c for c in df_out.columns if any(k in str(c) for k in ["付款方式", "取貨門市", "取貨日期", "發票"])]
    if cols_to_drop:
        df_out = df_out.drop(columns=cols_to_drop)
    df_out.columns = [str(c).replace("台", "臺") for c in df_out.columns]
    for col in df_out.columns:
        if col == col_phone_name or "電話" in col or "手機" in col:
            df_out[col] = df_out[col].apply(format_phone_number)
        else:
            df_out[col] = df_out[col].map(lambda x: str(x).replace("台", "臺") if isinstance(x, str) else x)
    return df_out

def highlight_duplicate_orders(workbook, sheet_name, df):
    if df.empty:
        return
    ws = workbook[sheet_name]
    name_col_idx = next((i for i, c in enumerate(df.columns) if "姓名" in str(c)), None)
    phone_col_idx = next((i for i, c in enumerate(df.columns) if "電話" in str(c) or "手機" in str(c)), None)
    if name_col_idx is None or phone_col_idx is None:
        return
    
    user_keys = [f"{str(row.iloc[name_col_idx]).strip()}_{str(row.iloc[phone_col_idx]).strip()}" for _, row in df.iterrows()]
    counts = pd.Series(user_keys).value_counts()
    dup_keys = counts[counts > 1].index.tolist()
    color_map = {k: COLOR_PALETTE[i % len(COLOR_PALETTE)] for i, k in enumerate(dup_keys)}
    
    for r_idx, k in enumerate(user_keys, start=2):
        if k in color_map:
            fill = PatternFill(start_color=color_map[k], end_color=color_map[k], fill_type="solid")
            for c_idx in range(1, len(df.columns) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = fill

# ==========================================
# 📋 TAB 1: 訂單自動分類整理
# ==========================================
with tab1:
    st.markdown("上傳原始訂單 Excel/CSV 後，系統會自動按物流方式分類並整理成專屬欄位！同姓名電話重複訂單自動塗上底色。")
    uploaded_file = st.file_uploader("請上傳原始訂單檔案 (支援 .xlsx, .xls, .csv)", type=["xlsx", "xls", "csv"], key="u1")

    if uploaded_file is not None:
        try:
            df = pd.read_csv(uploaded_file, dtype=str) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file, dtype=str)
            st.success(f"成功讀取檔案！共 {len(df)} 筆訂單。")
            st.markdown("---")
            
            cols = list(df.columns)
            c1, c2, c3 = st.columns(3)
            with c1:
                col_ship = st.selectbox("1. 物流方式欄位", cols, index=0 if "物流" not in cols else cols.index("物流"))
                col_name = st.selectbox("2. 收件人姓名欄位", cols, index=0 if "姓名" not in cols else cols.index("姓名"))
                col_email = st.selectbox("3. 收件人 Email 欄位", cols, index=0 if "Email" not in cols and "信箱" not in cols and "mail" not in cols else (cols.index("Email") if "Email" in cols else (cols.index("信箱") if "信箱" in cols else cols.index("mail"))))
            with c2:
                col_phone = st.selectbox("4. 收件人電話欄位", cols, index=0 if "電話" not in cols else cols.index("電話"))
                col_addr = st.selectbox("5. 收件地址欄位", cols, index=0 if "地址" not in cols else cols.index("地址"))
                col_items = st.selectbox("6. 購買品項欄位", cols, index=0 if "品項" not in cols and "商品" not in cols else (cols.index("品項") if "品項" in cols else cols.index("商品")))
            with c3:
                col_store = st.selectbox("7. 超商門市欄位", cols, index=0 if "門市" not in cols and "店" not in cols else (cols.index("門市") if "門市" in cols else cols.index("店")))

            if st.button("🚀 開始分類與整理", key="b1"):
                df_post = df[df[col_ship].astype(str).str.contains("郵局|宅配|快捷|包裹", na=False)].copy()
                df_711 = df[df[col_ship].astype(str).str.contains("711|7-11|交貨便|統一", na=False)].copy()
                df_familymart = df[df[col_ship].astype(str).str.contains("全家|店到店", na=False)].copy()
                df_sf_hk = df[df[col_ship].astype(str).str.contains("順豐|香港|SF", na=False)].copy()
                known_keywords = "郵局|宅配|快捷|包裹|711|7-11|交貨便|統一|全家|店到店|順豐|香港|SF"
                df_overseas = df[~df[col_ship].astype(str).str.contains(known_keywords, na=False)].copy()

                # 1. 整理郵局包裹 / 宅配
                post_data = []
                for _, row in df_post.iterrows():
                    c, d, r, rest = parse_taiwan_address(str(row.get(col_addr, '')))
                    post_data.append({
                        "收件人姓名": str(row.get(col_name, '')).replace("台", "臺"),
                        "收件人電話": format_phone_number(row.get(col_phone, '')),
                        "收件人Email": str(row.get(col_email, '')).strip(),
                        "購買品項": str(row.get(col_items, '')).strip(),
                        "完整地址": str(row.get(col_addr, '')).replace("台", "臺"),
                        "縣市": c, "鄉鎮市區": d, "路名": r, "剩餘地址": rest
                    })
                df_post_out = pd.DataFrame(post_data)

                # 2. 整理 7-11 店到店（店號在前，店名在後）
                seven_data = []
                for _, row in df_711.iterrows():
                    code, name = parse_store_info(str(row.get(col_store, '')))
                    seven_data.append({
                        "收件人姓名": str(row.get(col_name, '')).replace("台", "臺"),
                        "收件人電話": format_phone_number(row.get(col_phone, '')),
                        "收件人Email": str(row.get(col_email, '')).strip(),
                        "購買品項": str(row.get(col_items, '')).strip(),
                        "原始門市資訊": str(row.get(col_store, '')).replace("台", "臺"),
                        "收件店號": code,
                        "收件店名": name,
                        "寄件店號": "252975",
                        "寄件店名": "永吉門市"
                    })
                df_711_out = pd.DataFrame(seven_data)

                # 3. 整理 全家店到店（店號在前，店名在後）
                family_data = []
                for _, row in df_familymart.iterrows():
                    code, name = parse_store_info(str(row.get(col_store, '')))
                    family_data.append({
                        "收件人姓名": str(row.get(col_name, '')).replace("台", "臺"),
                        "收件人電話": format_phone_number(row.get(col_phone, '')),
                        "收件人Email": str(row.get(col_email, '')).strip(),
                        "購買品項": str(row.get(col_items, '')).strip(),
                        "原始門市資訊": str(row.get(col_store, '')).replace("台", "臺"),
                        "收件店號": code,
                        "收件店名": name,
                        "寄件店號": "024502",
                        "寄件店名": "永吉二店"
                    })
                df_familymart_out = pd.DataFrame(family_data)

                # 4. 香港順豐 & 海外
                df_sf_hk_out = clean_overseas_columns(df_sf_hk, col_phone)
                df_overseas_out = clean_overseas_columns(df_overseas, col_phone)

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_post_out.to_excel(writer, sheet_name='郵局包裹寄送', index=False)
                    df_711_out.to_excel(writer, sheet_name='711店到店', index=False)
                    df_familymart_out.to_excel(writer, sheet_name='全家店到店', index=False)
                    df_sf_hk_out.to_excel(writer, sheet_name='香港順豐到付', index=False)
                    df_overseas_out.to_excel(writer, sheet_name='其他海外寄送', index=False)
                    
                    wb = writer.book
                    highlight_duplicate_orders(wb, '郵局包裹寄送', df_post_out)
                    highlight_duplicate_orders(wb, '711店到店', df_711_out)
                    highlight_duplicate_orders(wb, '全家店到店', df_familymart_out)
                    highlight_duplicate_orders(wb, '香港順豐到付', df_sf_hk_out)
                    highlight_duplicate_orders(wb, '其他海外寄送', df_overseas_out)
                    
                output.seek(0)
                st.success("🎉 分類與欄位拆分完成！已補齊 Email、購買品項，並調整店號店名順序。")

                t1, t2, t3, t4, t5 = st.tabs(["郵局包裹", "7-11店到店", "全家店到店", "香港順豐", "其他海外"])
                with t1: st.dataframe(df_post_out)
                with t2: st.dataframe(df_711_out)
                with t3: st.dataframe(df_familymart_out)
                with t4: st.dataframe(df_sf_hk_out)
                with t5: st.dataframe(df_overseas_out)

                orig_name = uploaded_file.name.rsplit('.', 1)[0]
                st.download_button(
                    label="💾 下載分類整理後的 Excel 檔 (.xlsx)",
                    data=output,
                    file_name=f"{orig_name}_已分類出貨單.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"❌ 發生錯誤：{e}")

# ==========================================
# 🚚 TAB 2: 轉 7-11 批量匯入檔 (公版格式)
# ==========================================
with tab2:
    st.markdown("將整理好的 **711 收件資料 Excel** 上傳，系統會自動轉成 **7-11 官方交貨便批量匯入公版格式**！")
    
    col_sender1, col_sender2, col_sender3, col_sender4 = st.columns(4)
    with col_sender1:
        sender_name = st.text_input("寄件人姓名", value="有樂作業")
    with col_sender2:
        sender_phone = st.text_input("寄件人電話", value="0900000000")
    with col_sender3:
        sender_mail = st.text_input("寄件人 Mail", value="")
    with col_sender4:
        parcel_val = st.number_input("實際包裹價值", value=1000, step=100)

    uploaded_711_file = st.file_uploader("請上傳已整理好的 711 訂單 Excel 檔案", type=["xlsx", "xls", "csv"], key="u2")

    if uploaded_711_file is not None:
        try:
            df_711_in = pd.read_csv(uploaded_711_file, dtype=str) if uploaded_711_file.name.endswith('.csv') else pd.read_excel(uploaded_711_file, dtype=str)
            st.success(f"成功讀取 711 訂單，共 {len(df_711_in)} 筆。")
            
            cols_711 = list(df_711_in.columns)
            c7_1, c7_2, c7_3, c7_4 = st.columns(4)
            with c7_1:
                col_recv_name = st.selectbox("收件人姓名", cols_711, index=0 if "收件人姓名" not in cols_711 else cols_711.index("收件人姓名"))
            with c7_2:
                col_recv_phone = st.selectbox("收件人電話", cols_711, index=0 if "收件人電話" not in cols_711 else cols_711.index("收件人電話"))
            with c7_3:
                col_recv_email = st.selectbox("收件人 Email", cols_711, index=0 if "收件人Email" not in cols_711 and "Email" not in cols_711 else (cols_711.index("收件人Email") if "收件人Email" in cols_711 else cols_711.index("Email")))
            with c7_4:
                col_recv_store_code = st.selectbox("收件門市店號", cols_711, index=0 if "收件店號" not in cols_711 else cols_711.index("收件店號"))
            
            c7_5, _ = st.columns([1, 3])
            with c7_5:
                col_recv_store = st.selectbox("收件門市名稱", cols_711, index=0 if "收件店名" not in cols_711 else cols_711.index("收件店名"))

            if st.button("🚀 產生 7-11 批量匯入檔", key="b2"):
                header_row_1 = [
                    None, "寄件人姓名", "寄件人電話", "寄件人mail", "實際包裏價值",
                    "收件門市", "收件門市店號", "收件人姓名\n(請填寫證件姓名)", "收件人電話", "收件人mail",
                    "退貨門市\n(非必填，退貨門市未填寫者，則依原寄件門市為退貨門市。此欄位有填寫者，退貨門市店號必填)",
                    "退貨門市店號\n(非必填，退貨門市未填寫者，則依原寄件門市為退貨門市。此欄位有填寫者，退貨門市必填)"
                ]
                
                header_row_2 = [
                    "範例", "王小明", "0987654321", "xx@gmial.com", 1200,
                    "百建門市", "211147", "李大美", "0912345678", "oooooo@gmail.com",
                    "見晴門市", "217477"
                ]

                data_rows = [header_row_1, header_row_2]

                for _, row in df_711_in.iterrows():
                    r_name = str(row.get(col_recv_name, '')).strip()
                    r_phone = format_phone_number(row.get(col_recv_phone, ''))
                    r_email = str(row.get(col_recv_email, '')).strip()
                    r_store = str(row.get(col_recv_store, '')).strip()
                    r_store_code = str(row.get(col_recv_store_code, '')).strip()

                    data_rows.append([
                        None,             # A欄 (留空)
                        sender_name,      # 寄件人姓名
                        sender_phone,     # 寄件人電話
                        sender_mail,      # 寄件人mail
                        parcel_val,       # 實際包裹價值
                        r_store,          # 收件門市
                        r_store_code,     # 收件門市店號
                        r_name,           # 收件人姓名
                        r_phone,          # 收件人電話
                        r_email,          # 收件人mail
                        "",               # 退貨門市
                        ""                # 退貨門市店號
                    ])

                wb_711 = Workbook()
                ws_711 = wb_711.active
                ws_711.title = "工作表1"

                for row in data_rows:
                    ws_711.append(row)

                ws_711.cell(row=1, column=1, value=None)

                output_711 = io.BytesIO()
                wb_711.save(output_711)
                output_711.seek(0)

                st.success("🎉 7-11 批量匯入檔已完美生成！收件人 Email 已自動帶入。")
                
                preview_df = pd.DataFrame(data_rows[2:], columns=[
                    "未命名", "寄件人姓名", "寄件人電話", "寄件人mail", "實際包裹價值",
                    "收件門市", "收件門市店號", "收件人姓名", "收件人電話", "收件人mail", "退貨門市", "退貨門市店號"
                ])
                st.dataframe(preview_df.drop(columns=["未命名"]))

                st.download_button(
                    label="💾 下載 7-11 官方批量匯入 Excel (.xlsx)",
                    data=output_711,
                    file_name="一般交貨便-取貨不付款_匯入檔.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"❌ 處理 711 匯入檔時發生錯誤：{e}")
